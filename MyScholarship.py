from ast import Not
from contextlib import nullcontext
from bs4 import BeautifulSoup
import openpyxl 
from openpyxl import load_workbook
import requests

class MyScholarships:
    MyScholarshipsLink = "https://dallascollege.academicworks.com/opportunities"        
        
    def get_soup(link=""):
        req = requests.get(link)
        soup = BeautifulSoup(req.text, "html.parser")
        return soup

    @staticmethod
    def _get_pages():
        pages = []
        soup = MyScholarships.get_soup(MyScholarships.MyScholarshipsLink)
        for p in soup.find_all('option',attrs={'name':'page'}):
            pages.append(p['data-direct-url'])
        return pages
    
    @staticmethod
    def _get_name(tag):
        try:
            return tag.find("a").text
        except:
            print("<a> link tag not present")
            return 0

    @staticmethod
    def _get_link(tag):
        try:
            return MyScholarships.MyScholarshipsLink + tag.find("a")["href"].replace('opportunities/','')
        except:
            print("<a> link tag not present")
            return 0

    @staticmethod
    def _get_award(tag):
        try:
            for td in tag.findAll('td')[:1]:
                return td.text.strip()
        except:
            print("<td> tag with award not found")
            return 0

    @staticmethod
    def _get_deadline(tag):
        try:
            for td in tag.findAll('td')[1:]:
                return td.text.strip()
        except:
            print("<td> tag with award not found")
            return 0

    @staticmethod
    def _get_questions(lnk):
        q = []
        try:
            soup = MyScholarships.get_soup(lnk)
        except:
            print(f'Invalid Link')
        else:
            for l in soup.find_all(class_="js-question"):
                q.append(l.text)
        return q

    @staticmethod
    def _get_scholarship():
        MyScholarshipsList = []
        for page in MyScholarships._get_pages():
            soup  = MyScholarships.get_soup(MyScholarships.MyScholarshipsLink + page)
            for tr in soup.findAll("tr")[1:]:
                dic = {
                "Deadline" : MyScholarships._get_deadline(tr),
                "Award"    : MyScholarships._get_award(tr),
                "Name"     : MyScholarships._get_name(tr),
                "Link"     : MyScholarships._get_link(tr),
                "Questions": MyScholarships._get_questions(MyScholarships._get_link(tr))
                }
                if dic["Deadline"] == "Ended" or dic["Deadline"] == "":
                    continue
                else:
                    MyScholarshipsList.append(dic)
        return MyScholarshipsList

    def scholarship_xlsx(self,filename = "Scholarships.xlsx"):
        wb = openpyxl.Workbook() 
        sheet = wb.active
        _row = 1
        for dic in MyScholarships._get_scholarship():
            sheet.cell(row=_row , column=1).value = dic["Deadline"]
            sheet.cell(row=_row , column=2).value = dic["Award"]
            sheet.cell(row=_row , column=3).value = dic["Name"]
            sheet.cell(row=_row , column=4).value = dic["Link"]
            sheet.cell(row=_row ,column=5).value = str(dic["Questions"])
            if len(dic["Questions"]) != 0:
                for q in dic["Questions"]:
                    _row += 1
                    sheet.cell(row=_row , column=1).value = dic["Deadline"]
                    sheet.cell(row=_row , column=2).value = dic["Award"]
                    sheet.cell(row=_row , column=3).value = dic["Name"]
                    sheet.cell(row=_row , column=4).value = dic["Link"]
                    sheet.cell(row=_row , column=5).value = q
                _row += 1
            elif len(dic["Questions"]) == 0:
                sheet.cell(row = _row, column=5).value = "No Questions"
                _row += 1
        wb.save(filename)
        return filename
    

    def get_list_length(self,filename):
        wb = load_workbook(filename)
        sheet = wb.active
        i = 1
        while sheet.cell(row=i,column=1).value != None:
            i += 1
        return i-1
    
s = MyScholarships()
print(s.scholarship_xlsx() + "file was made")

  